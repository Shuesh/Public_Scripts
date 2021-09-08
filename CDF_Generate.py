#imports to read from Excel file
from openpyxl import load_workbook


#get file path
file_path = input("Enter full path to file: ")

wb = load_workbook(file_path)
ws = wb.active

colA = ws['A']
numRows = len(colA)
print(numRows)

row1 = ws[1]
numCols = len(row1)
print(numCols)

value_array = []
second_harmonic_first = False
third_harmonic_first = False
missing_second = True
missing_third = True

for row in ws.iter_rows(min_row=1, max_col=9, max_row=numRows):
    for cell in row:
        str(value_array.append(cell.value))

for val in value_array[0:numCols-1]:
    if "Second" in val or "second" in val:
        second_harmonic_first = True
        missing_second = False
    if "Third" in val or "third" in val:
        third_harmonic_first = True
        missing_third = False

    if ("Second" in val or "second" in val) and third_harmonic_first:
        missing_second = False
        break
    if ("Third" in val or "third" in val) and second_harmonic_first:
        missing_third = False
        break

#Determined the order of second/third harmonic table
#Next we need to build up dictionaries to associate values

# second_element_dict = {:}
# second_marconi_dict = {:}
# third_element_dict = {:}
# third_marconi_dict = {:}

